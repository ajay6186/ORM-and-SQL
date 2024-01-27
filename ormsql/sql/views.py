from django.shortcuts import render
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django.db import connection
from django.db import models
import openpyxl


@api_view(['GET'])
def select_all_p(request):
    with connection.cursor() as cursor:
        cursor.execute("SELECT * FROM employees")
        results = cursor.fetchall()
        data_list = []
    for result in results:
        data = {
            'employee_id': result[0],
            'last_name': result[1],
            'first_name': result[2],
            'title': result[3],
            'title_of_courtesy': result[4],
            'birth_date': result[5],
            'hire_date': result[6],
            'address': result[7],
            'city': result[8],
            'region': result[9],
            'postal_code': result[10],
            'country': result[11],
            'home_phone': result[12],
            'extension': result[13],
            'photo': result[14],
            'notes': result[15],
            'reports_to': result[16],
            'photo_path': result[17],
        }
        data_list.append(data)
    return Response(data_list)



@api_view(['GET'])
def select_few_columns_p(request):
      with connection.cursor() as cursor:
        cursor.execute("SELECT first_name, last_name, address, country from employees")
        results = cursor.fetchall()
        data_list = []
      for result in results:
            data = {
                'first_name': result[0],
                'last_name': result[1],
                'address': result[2],
                'country': result[3],
            }
            data_list.append(data)
      return Response(data_list)


from django.http import JsonResponse

@api_view(['GET'])
def select_all(request):
    data = {'message': 'This response from SQL'}
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM employees")
    result = cursor.fetchall()
    response = []
    for r in result:
        d = {}
        d['employee_id'] =    r[0]
        d['last_name'] =    r[1]
        d['first_name'] =    r[2]
        d['title'] =    r[3]
        d['title_of_courtesy'] =    r[4]
        d['birth_date'] =    r[5].strftime('%Y-%m-%d')
        d['hire_date'] =    r[6].strftime('%Y-%m-%d')
        d['address'] =    r[7]
        d['city'] =    r[8]
        d['region'] =    r[9]
        d['postal_code'] =    r[10]
        d['country'] =    r[11]
        d['home_phone'] =    r[12]
        d['extension'] =    r[13]
        d['photo'] =    r[14].tobytes().decode('utf-8')  if isinstance(r[14], str) else None,
        d['notes'] =    r[15]
        d['reports_to'] =    r[16] if r[16] else None
        d['photo_path'] =    r[17]
        response.append(d)
    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)

@api_view(['GET'])
def select_few_columns(request):
    data = {'message': 'This response from SQL'}
    cursor = connection.cursor()
    cursor.execute("SELECT first_name, last_name, address, country FROM employees")
    result = cursor.fetchall()
    response = []
    for r in result:
        d = {}
        d['first_name'] =    r[0]
        d['last_name'] =    r[1]
        d['address'] =    r[2]
        d['country'] =    r[3]
        response.append(d)
    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)


@api_view(['GET'])
def Old_price_New_price_p(request):
      with connection.cursor() as cursor:
        cursor.execute("select  product_name, unit_price as old_price, unit_price*2 as new_price  from products")
        results = cursor.fetchall()
        data_list = []
      for result in results:
            data = {
                'product_name': result[0],
                'old_price': result[1],
                'unit_price': result[2],
               
            }
            data_list.append(data)
      return Response(data_list)


# SELECT PRODUCT_NAME,

# UNIT_PRICE AS Old_Price,

# UNIT_PRICE*2 AS new_Price

# FROM PRODUCTS;
@api_view(['GET'])
def select_old_price_and_new_price(request):
    data = {'message': 'This response from SQL'}
    cursor = connection.cursor()
    cursor.execute("SELECT PRODUCT_NAME, UNIT_PRICE AS Old_Price, UNIT_PRICE*2 AS new_Price FROM products")
    result = cursor.fetchall()
    response = []
    for r in result:
        d = {}
        d['product_name'] =    r[0]
        d['Old_Price'] =    r[1]
        d['new_Price'] =    r[2]
        response.append(d)
    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)


@api_view(['GET'])
def select_report_with_comment_p(request):
    cursor = connection.cursor()
    cursor.execute("SELECT 'Our customer=>' AS COMMENT,* FROM CUSTOMERS")
    data = cursor.fetchall()
    response = []
    for result in data:
        d = {
            'comment'      : result[0],
            'customer_id'  : result[1],
            'company_name' : result[2],
            'contact_name' : result[3],
            'contact_title': result[4],
            'address'      : result[5],
            'city'         : result[6],
            'region'       : result[7],
            'postal_code'  : result[8],
            'country'      : result[9],
            'phone'        : result[10],
            'fax'          :  result[11],
        }
        response.append(d)
    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)


def select_distinct_p(request):
    cursor = connection.cursor()
    cursor.execute("SELECT DISTINCT COUNTRY FROM CUSTOMERS") # remove the duplicate element
    data = cursor.fetchall()
    response = []
    for result in data:
        d = {
            'country' : result[0],
        }
        response.append(d)
    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)


def select_distinct_p_list(request):
    cursor = connection.cursor()
    cursor.execute("SELECT DISTINCT COUNTRY FROM CUSTOMERS") # remove the duplicate element
    data = cursor.fetchall()
    d={}
    d['country'] = []
    for result in data:
        data= result[0]
        d['country'].append(data)
    return JsonResponse(d, safe=False, status=status.HTTP_200_OK)


@api_view(['GET'])
def alias_sql_alias_p(request):  # change the coloumn name using alising
    cursor = connection.cursor()
    cursor.execute("SELECT EMPLOYEE_ID AS ID, FIRST_NAME AS NAME, TITLE_OF_COURTESY AS TITLE FROM EMPLOYEES") # remove the duplicate element
    data = cursor.fetchall()
    response = []
    for result in data:
        d = {
            'ID' : result[0],
            'NAME' : result[1],
            'TITLE FROM EMPLOYEES' : result[2],
        }
        response.append(d)
  
    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)


@api_view(['GET'])
def reporting_with_Concatenation_p(request):
    cursor = connection.cursor()
    cursor.execute("SELECT CONCAT(TITLE_OF_COURTESY,'',FIRST_NAME) AS NAME FROM EMPLOYEES") # same work both query
    # cursor.execute("SELECT TITLE_OF_COURTESY||' '||FIRST_NAME AS NAME FROM EMPLOYEES")
    data = cursor.fetchall()
    d={}
    d['NAME'] = []
    for result in data:
        data= result[0]
        d['NAME'].append(data)
    return JsonResponse(d, safe=False, status=status.HTTP_200_OK)


@api_view(['GET'])
def reporting_with_Concatenation_p_New(request):
    cursor = connection.cursor()
    cursor.execute("SELECT TITLE_OF_COURTESY||' '||FIRST_NAME||' '||LAST_NAME AS NAME FROM EMPLOYEES")
    data = cursor.fetchall()
    d={}
    d['NAME'] = []
    for result in data:
        data= result[0]
        d['NAME'].append(data)
    return JsonResponse(d, safe=False, status=status.HTTP_200_OK)


@api_view(['GET'])
def  reporting_with_concatenation_customer_location_p(request):  # change the coloumn name using alising
    cursor = connection.cursor()
    cursor.execute("SELECT COMPANY_NAME, CONCAT('is from ', city, ' in ', country) AS LOCATION FROM CUSTOMERS") 
    data = cursor.fetchall()
    response = []
    for result in data:
        d = {
            'COMPANY_NAME' : result[0],
            'LOCATION' : result[1],
        }
        response.append(d)
    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)


@api_view(['GET'])
def  reporting_with_concatenation_customer_location_New_p(request): 
    cursor = connection.cursor()
    cursor.execute("SELECT customer_id,COMPANY_NAME || ' is from '|| city|| ' in '|| country AS LOCATION FROM CUSTOMERS")
    data = cursor.fetchall()
    d={}
    d['LOCATION']=[]
    for result in data:
        data=result[0]
        data=result[1]
        d['LOCATION'].append(data)
    return JsonResponse(d, safe=False, status=status.HTTP_200_OK)


def reporting_with_Concatenation_p(request):
    cursor = connection.cursor()
    cursor.execute("SELECT PRODUCT_NAME AS PRODUCT, CONCAT('$',UNIT_PRICE * UNITS_IN_STOCK) AS VALUE FROM PRODUCTS")
    data = cursor.fetchall()
    response = []
    for result in data:
        d = {
            'PRODUCT' : result[0],
            'VALUE' : result[1],
        }
        response.append(d)
    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)


def report_formatting_without_concatention_p(request):  #without cocatention 
    cursor = connection.cursor()
    cursor.execute("SELECT FORMAT('%s  %s',LAST_NAME,FIRST_NAME) FULL_NAME FROM EMPLOYEES")
    data = cursor.fetchall()
    d={}
    d['FULL_NAME']=[]
    for result in data:
        l=result[0]
        d['FULL_NAME'].append(l)
    return JsonResponse(d, safe=False, status=status.HTTP_200_OK)


def report_formatting_add_two_coloumn_in_one_coloumn_p(request):
    cursor = connection.cursor()
    cursor.execute("SELECT FORMAT('FirstName:  %-10s LastName: %-10s', FIRST_NAME,LAST_NAME) full_name FROM EMPLOYEES")
    data = cursor.fetchall()
    d={}
    d['full_name']=[]
    for result in data:
        l=result[0]
        d['full_name'].append(l)
    return JsonResponse(d, safe=False, status=status.HTTP_200_OK)


@api_view(['GET'])
def report_formatting_add_three_coloumn_in_one_coloumn_p(request):
    cursor = connection.cursor()
    cursor.execute("SELECT FORMAT('%s %s  %s', TITLE_OF_COURTESY,LAST_NAME,FIRST_NAME) full_name FROM EMPLOYEES")
    data = cursor.fetchall()
    d={}
    d['full_name']=[]
    for result in data:
        l=result[0]
        d['full_name'].append(l)
    return JsonResponse(d, safe=False, status=status.HTTP_200_OK)


def round_amount_two_decimal(request): 
    cursor = connection.cursor()
    cursor.execute("SELECT ROUND(UNIT_PRICE::numeric,2)FROM PRODUCTS")
    data = cursor.fetchall()
    d={}
    d['ROUND']=[]
    for result in data:
        l=result[0]
        d['ROUND'].append(l)
    return JsonResponse(d, safe=False, status=status.HTTP_200_OK)


def round_amount_one_decimal(request): 
    cursor = connection.cursor()
    cursor.execute("SELECT ROUND(UNIT_PRICE::numeric,1)FROM PRODUCTS")
    data = cursor.fetchall()
    d={}
    d['ROUND']=[]
    for result in data:
        l=result[0]
        d['ROUND'].append(l)
    return JsonResponse(d, safe=False, status=status.HTTP_200_OK)


def round_amounformattiong_53(request):
    cursor = connection.cursor()
    cursor.execute("SELECT PRODUCT_NAME AS PRODUCT,CONCAT('$', ROUND(CAST(unit_price * units_in_stock as numeric), 2)) as Value FROM PRODUCTS")
    data = cursor.fetchall()
    response = []
    for result in data:
        d = {
            'PRODUCT' : result[0],
            'numeric' : result[1],
        }
        response.append(d)
    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)


def order_by_company_name(request):
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM CUSTOMERS ORDER BY COMPANY_NAME")
    data = cursor.fetchall()
    response = []
    for result in data:
        d = {
            'customer_id' : result[0],
            'company_name' : result[1],
            'contact_name' : result[2],
            'contact_title' : result[3],
            'address' : result[4],
            'city' : result[5],
            'region' : result[6],
            'postal_code' : result[7],
            'country' : result[8],
            'phone' : result[9],
            'fax' : result[10],
        }
        response.append(d)
    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)




def order_by_customer_id_asc_order_id_desc(request):
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM orders ORDER BY customer_id ASC, order_id DESC")
    data = cursor.fetchall()
    response = []
    for result in data:
        d = {
            'order_id' : result[0],
            'customer_id' : result[1],
            'employee_id' : result[2],
            'order_date' : result[3],
            'required_date' : result[4],
            'shipped_date' : result[5],
            'ship_via' : result[6],
            'freight' : result[7],
            'ship_name' : result[8],
            'ship_address' : result[9],
            'ship_city' : result[10],
            'ship_region' : result[11],
            'ship_postal_code ' : result[12],
            'ship_country' : result[13],
        }
        response.append(d)
    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)



def order_by_customer_id_asc_order_id_asc(request):
    cursor = connection.cursor()
    cursor.execute("SELECT * FROM orders ORDER BY customer_id ASC, order_id ASC")
    data = cursor.fetchall()
    response = []
    for result in data:
        d = {
            'order_id' : result[0],
            'customer_id' : result[1],
            'employee_id' : result[2],
            'order_date' : result[3],
            'required_date' : result[4],
            'shipped_date' : result[5],
            'ship_via' : result[6],
            'freight' : result[7],
            'ship_name' : result[8],
            'ship_address' : result[9],
            'ship_city' : result[10],
            'ship_region' : result[11],
            'ship_postal_code ' : result[12],
            'ship_country' : result[13],
        }
        response.append(d)
    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)


@api_view(['POST'])
def where_sales_manager_details(request):
    title=request.POST.get('title')
    cursor = connection.cursor()
    # cursor.execute("SELECT * FROM EMPLOYEES WHERE title = %s", [title])
    query=F"SELECT * FROM EMPLOYEES WHERE title = '{title}'"
    cursor.execute(query) 
    data = cursor.fetchall()
    data_list = []
    for result in data:
        data = {
            'employee_id': result[0],
            'last_name': result[1],
            'first_name': result[2],
            'title': result[3],
            'title_of_courtesy': result[4],
            'birth_date': result[5],
            'hire_date': result[6],
            'address': result[7],
            'city': result[8],
            'region': result[9],
            'postal_code': result[10],
            'country': result[11],
            'home_phone': result[12],
            'extension': result[13],
            'photo': result[14],
            'notes': result[15],
            'reports_to': result[16],
            'photo_path': result[17],
        }
        data_list.append(data)
    return Response(data_list)

@api_view(['POST'])
def where_owner_details(request):
    contact_title = request.POST.get('contact_title')
    cursor = connection.cursor()
    query=F"SELECT company_name, contact_title, address, city, country, phone FROM customers WHERE contact_name = {contact_title}" 
    cursor.execute(query) 
    data = cursor.fetchall()
    response = []
    for result in data:
        d = {
            'company_name' : result[0],
            'contact_title' : result[1],
            'address' : result[2],
            'city' : result[3],
            'country' : result[4],
            'phone' : result[5],
        }
        response.append(d)
    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)

@api_view(['POST'])
def where_specific_order_id(request):
    order_id = request.POST.get('order_id')
    cursor = connection.cursor()
    query=F"SELECT * FROM orders WHERE order_id  = {order_id}" 
    cursor.execute(query) 
    data = cursor.fetchall()
    response = []
    for result in data:
        d = {
            'order_id' : result[0],
            'customer_id' : result[1],
            'employee_id' : result[2],
            'order_date' : result[3],
            'required_date' : result[4],
            'shipped_date' : result[5],
            'ship_via' : result[6],
            'freight' : result[7],
            'ship_name' : result[8],
            'ship_address' : result[9],
            'ship_city' : result[10],
            'ship_region' : result[11],
            'ship_postal_code ' : result[12],
            'ship_country' : result[13],
        }
        response.append(d)
    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)


@api_view(['POST'])
def where_searching_product(request):
    unit_price = request.POST.get('unit_price')
    cursor = connection.cursor()
    query= F"SELECT product_name, unit_price FROM PRODUCTS WHERE unit_price  = {unit_price}" 
    cursor.execute(query) 
    data = cursor.fetchall()
    response = []
    for result in data:
        d = {
            'Prodcut_name' : result[0],
            'Unit_price' : result[1],
        }
        response.append(d)
    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)


@api_view(['POST'])
def where_searching_for_customer_id(request):
    customer_id = request.POST.get('customer_id')
    cursor = connection.cursor()
    query=F"SELECT customer_id, company_name FROM Customers WHERE customer_id = {customer_id}"
    cursor.execute(query) 
    data = cursor.fetchall()
    response = []
    for result in data:
        d = {
            'customer_id' : result[0],
            'company_name' : result[1],
          }
        response.append(d)
    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)


@api_view(['POST'])
def where_product_above(request):
    unit_price = request.POST.get('unit_price')
    cursor = connection.cursor()
    query=F"SELECT product_name,unit_price FROM products WHERE unit_price > '{unit_price}' ORDER BY unit_price"
    cursor.execute(query)
    # F"SELECT * FROM EMPLOYEES WHERE title = '{title}'"
    data = cursor.fetchall()
    response = []
    for result in data:
        d = {
            'product_name' : result[0],
            'unit_price' : result[1],
          }
        response.append(d)
    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)

@api_view(['POST'])
def where_product_less(request):
    unit_price = request.POST.get('unit_price')
    cursor = connection.cursor()
    query=F"SELECT product_name,unit_price FROM products WHERE unit_price < '{unit_price}' ORDER BY unit_price"
    cursor.execute(query)
    # F"SELECT * FROM EMPLOYEES WHERE title = '{title}'"
    data = cursor.fetchall()
    response = []
    for result in data:
        d = {
            'product_name' : result[0],
            'unit_price' : result[1],
          }
        response.append(d)
    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)

@api_view(['POST'])
def where_product_whose_price_are_40_above(request):
    unit_price = request.POST.get('unit_price')
    cursor = connection.cursor()
    query=F"SELECT product_name,unit_price FROM products WHERE unit_price >= '{unit_price}' ORDER BY unit_price"
    cursor.execute(query)
    # F"SELECT * FROM EMPLOYEES WHERE title = '{title}'"
    data = cursor.fetchall()
    response = []
    for result in data:
        d = {
            'product_name' : result[0],
            'unit_price' : result[1],
          }
        response.append(d)
    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)

@api_view(['POST'])
def where_product_with_range(request):
    order_id_1 = request.POST.get('order_id_1')
    order_id_2 = request.POST.get('order_id_2')
    cursor = connection.cursor()
    query=F"SELECT * FROM orders WHERE order_id < {order_id_1} or order_id> {order_id_2}"
    cursor.execute(query)
    # F"SELECT * FROM EMPLOYEES WHERE title = '{title}'"
    data = cursor.fetchall()
    response = []
    for result in data:
        d = {
            'order_id' : result[0],
            'customer_id' : result[1],
            'employee_id' : result[2],
            'order_date' : result[3],
            'required_date' : result[4],
            'shipped_date' : result[5],
            'ship_via' : result[6],
            'freight' : result[7],
            'ship_name' : result[8],
            'ship_address' : result[9],
            'ship_city' : result[10],
            'ship_region' : result[11],
            'ship_postal_code ' : result[12],
            'ship_country' : result[13],
        }
        response.append(d)
    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)

@api_view(['POST'])
def where_product_rostocking(request):
    units_in_stock = request.POST.get('units_in_stock')
    reorder_level = request.POST.get('reorder_level')
    cursor = connection.cursor()
    query = F"SELECT product_name, units_in_stock, reorder_level FROM products WHERE units_in_stock < {units_in_stock} AND reorder_level > {reorder_level} ORDER BY units_in_stock"
    cursor.execute(query)
    data = cursor.fetchall()
    response = []
    for result in data:
        d = {
            'product_name': result[0],
            'units_in_stock': result[1],
            'reorder_level': result[2],
        }
        response.append(d)

    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)

@api_view(['POST'])
def order_amount_of_10540_or_above(request):
    unit_price = request.POST.get('unit_price')
    quantity = request.POST.get('quantity')
    cursor = connection.cursor()
    query = F"SELECT order_id, unit_price * quantity AS Amount FROM order_details WHERE unit_price * quantity >= {float(quantity) * float(unit_price)}ORDER BY Amount DESC"
    cursor.execute(query)
    data = cursor.fetchall()
    response = []
    for result in data:
        d = {
            'order_id': result[0],
            'Amount': result[1],
        }
        response.append(d)

    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)

@api_view(['POST'])
def where_not_mr(request):
    title_of_courtesy = request.POST.get('title_of_courtesy')
    cursor = connection.cursor()
    query= F"SELECT * FROM employees WHERE title_of_courtesy !={title_of_courtesy}"
    cursor.execute(query)
    datas = cursor.fetchall()
    response = []
    for result in datas:
        data = {
            'employee_id': result[0],
            'last_name': result[1],
            'first_name': result[2],
            'title': result[3],
            'title_of_courtesy': result[4],
            'birth_date': result[5],
            'hire_date': result[6],
            'address': result[7],
            'city': result[8],
            'region': result[9],
            'postal_code': result[10],
            'country': result[11],
            'home_phone': result[12],
            'extension': result[13],
            'photo': result[14].tobytes().decode('utf-8')  if isinstance(result[14], str) else None,
            'notes': result[15],
            'reports_to': result[16],
            'photo_path': result[17],
        }
        response.append(data)

    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)

@api_view(['POST'])
def where_not_sales_representative(request):
    title = request.POST.get('title')
    cursor = connection.cursor()
    query = F"SELECT * FROM employees WHERE title != {title}"
    cursor.execute(query)
    datas = cursor.fetchall()
    response = []
    for result in datas:
        data = {
            'employee_id': result[0],
            'last_name': result[1],
            'first_name': result[2],
            'title': result[3],
            'title_of_courtesy': result[4],
            'birth_date': result[5],
            'hire_date': result[6],
            'address': result[7],
            'city': result[8],
            'region': result[9],
            'postal_code': result[10],
            'country': result[11],
            'home_phone': result[12],
            'extension': result[13],
            'photo': result[14].tobytes().decode('utf-8')  if isinstance(result[14], str) else None,
            'notes': result[15],
            'reports_to': result[16],
            'photo_path': result[17],
        }
        response.append(data)

    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)

@api_view(['POST'])
def where_price_between_two_figure(request):
    min_price = request.POST.get('min_price')
    max_price = request.POST.get('max_price')
    cursor = connection.cursor()
    query = F"SELECT PRODUCT_NAME, UNIT_PRICE FROM PRODUCTS WHERE UNIT_PRICE >= {min_price} AND UNIT_PRICE <= {max_price} ORDER BY UNIT_PRICE"
    cursor.execute(query)
    data = cursor.fetchall()
    response = []
    for result in data:
        d = {
            'product_name': result[0],
            'units_price': result[1],
          
        }
        response.append(d)

    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)

@api_view(['GET'])
def where_price_not_between_two_figure(request):
    min_price = request.POST.get('min_price')
    max_price = request.POST.get('max_price')
    cursor = connection.cursor()
    query = F"SELECT PRODUCT_NAME, UNIT_PRICE FROM PRODUCTS WHERE UNIT_PRICE NOT BETWEEN {min_price} AND  {max_price} ORDER BY UNIT_PRICE"
    cursor.execute(query)
    data = cursor.fetchall()
    response = []
    for result in data:
        d = {
            'product_name': result[0],
            'units_price': result[1],
          
        }
        response.append(d)

    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)




# Create a new workbook and select the active sheet
workbook = openpyxl.Workbook()
sheet = workbook.active
@api_view(['POST'])
def where_two_between(request):
    min_price = request.POST.get('min_price')
    max_price = request.POST.get('max_price')
    min_price_1 = request.POST.get('min_price_1')
    max_price_1 = request.POST.get('max_price_1')

    # min_price = request.data.get('min_price')
    # max_price = request.data.get('max_price')
    # min_price_1 = request.data.get('min_price_1')
    # max_price_1 = request.data.get('max_price_1')

    cursor = connection.cursor()
    query = F"SELECT PRODUCT_NAME, UNIT_PRICE FROM PRODUCTS WHERE UNIT_PRICE BETWEEN {min_price} AND {max_price} OR UNIT_PRICE BETWEEN {min_price_1} AND {max_price_1} ORDER BY UNIT_PRICE"
    cursor.execute(query)
    data = cursor.fetchall()
    response = []

   ################################################################################
    
    sheet['A1'] = 'product_name'
    sheet['B1'] = 'units_price'

    # Starting row index for data
    start_row = 2

    # Populate the columns with data
    for result in data:
        sheet.cell(row=start_row, column=1, value=result[0])  # Product Name column
        sheet.cell(row=start_row, column=2, value=result[1])  # Units price column
        start_row += 1

    # Save the workbook
    workbook.save('Product_Unit_prices.xlsx')

   #################################################################################

    for result in data:
        d = {
            'product_name': result[0],
            'units_price': result[1],
        }
        response.append(d)

    return JsonResponse(response, safe=False, status=status.HTTP_200_OK)



@api_view(['GET'])
def where_customer_from_contain_city(request):
    city=request.POST.get('city')
    cursor = connection.cursor()
    query= F"SELECT * FROM CUSTOMERS WHERE CITY IN ('{city}')"
    cursor.execute(query) 
    data = cursor.fetchall()
    data_list = []
    for result in data:
        data = {
            'employee_id': result[0],
            'last_name': result[1],
            'first_name': result[2],
            'title': result[3],
            'title_of_courtesy': result[4],
            'birth_date': result[5],
            'hire_date': result[6],
            'address': result[7],
            'city': result[8],
            'region': result[9],
            'postal_code': result[10],
            'country': result[11],
            'home_phone': result[12],
            'extension': result[13],
            'photo': result[14],
            'notes': result[15],
            'reports_to': result[16],
            'photo_path': result[17],
        }
        data_list.append(data)
    return JsonResponse(data_list, safe=False, status=status.HTTP_200_OK)